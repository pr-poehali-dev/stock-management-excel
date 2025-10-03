"""
Business: Экспорт данных товаров из БД в Excel файл
Args: event - dict с httpMethod
      context - объект с request_id, function_name
Returns: Excel файл в base64 или JSON с ошибкой
"""

import json
import os
import base64
from typing import Dict, Any
from io import BytesIO
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'GET')
    
    # Handle CORS OPTIONS request
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }
    
    if method != 'GET':
        return {
            'statusCode': 405,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({'error': 'Method not allowed'})
        }
    
    database_url = os.environ.get('DATABASE_URL')
    if not database_url:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({'error': 'DATABASE_URL not configured'})
        }
    
    try:
        conn = psycopg2.connect(database_url)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT name, sku, quantity, min_stock, price, batch, created_at, updated_at
            FROM t_p72161094_stock_management_exc.products
            ORDER BY name
        """)
        
        products = cursor.fetchall()
        cursor.close()
        conn.close()
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({'error': f'Database error: {str(e)}'})
        }
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Название", "Артикул", "Количество", "Мин. остаток", "Цена (₽)", "Партия", "Создан", "Обновлен"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for product in products:
        name, sku, quantity, min_stock, price, batch, created_at, updated_at = product
        ws.append([
            name,
            sku,
            quantity,
            min_stock,
            float(price) if price else 0,
            batch or "",
            created_at.strftime("%Y-%m-%d %H:%M") if created_at else "",
            updated_at.strftime("%Y-%m-%d %H:%M") if updated_at else ""
        ])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    excel_base64 = base64.b64encode(excel_file.read()).decode('utf-8')
    
    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename="stock_products.xlsx"',
            'Access-Control-Allow-Origin': '*'
        },
        'isBase64Encoded': True,
        'body': excel_base64
    }