"""
Business: Импорт товаров из Excel файла в БД
Args: event - dict с httpMethod, body (base64 Excel файл)
      context - объект с request_id, function_name
Returns: JSON с результатом импорта
"""

import json
import os
import base64
from typing import Dict, Any, List
from io import BytesIO
import psycopg2
from openpyxl import load_workbook
from decimal import Decimal


def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'POST')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }
    
    if method != 'POST':
        return {
            'statusCode': 405,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({'error': 'Method not allowed'})
        }
    
    database_url = os.environ.get('DATABASE_URL')
    if not database_url:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({'error': 'DATABASE_URL not configured'})
        }
    
    body_str = event.get('body', '{}')
    if not body_str or body_str == '{}':
        return {
            'statusCode': 400,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({'error': 'No file provided'})
        }
    
    try:
        body_data = json.loads(body_str)
        file_base64 = body_data.get('file')
        
        if not file_base64:
            return {
                'statusCode': 400,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'isBase64Encoded': False,
                'body': json.dumps({'error': 'No file provided'})
            }
        
        file_bytes = base64.b64decode(file_base64)
        excel_file = BytesIO(file_bytes)
        
        wb = load_workbook(excel_file)
        ws = wb.active
        
        products: List[Dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            name = str(row[0]) if row[0] else ''
            inventory_number = str(row[1]) if row[1] else ''
            quantity = float(row[2]) if row[2] else 0.0
            unit = str(row[3]) if row[3] else 'шт'
            min_stock = float(row[4]) if row[4] else 0.0
            price = float(row[5]) if row[5] else 0.0
            batch = str(row[6]) if row[6] else ''
            
            products.append({
                'name': name,
                'inventory_number': inventory_number,
                'quantity': quantity,
                'unit': unit,
                'min_stock': min_stock,
                'price': price,
                'batch': batch
            })
        
        conn = psycopg2.connect(database_url)
        cursor = conn.cursor()
        
        inserted = 0
        updated = 0
        
        for product in products:
            safe_inv = str(product['inventory_number']).replace("'", "''")
            safe_name = str(product['name']).replace("'", "''")
            safe_unit = str(product['unit']).replace("'", "''")
            safe_batch = str(product['batch']).replace("'", "''")
            
            cursor.execute(
                f"SELECT id FROM products WHERE inventory_number = '{safe_inv}'"
            )
            
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute(
                    f"UPDATE products SET name = '{safe_name}', quantity = {product['quantity']}, " +
                    f"unit = '{safe_unit}', min_stock = {product['min_stock']}, price = {product['price']}, " +
                    f"batch = '{safe_batch}', updated_at = NOW() WHERE inventory_number = '{safe_inv}'"
                )
                updated += 1
            else:
                cursor.execute(
                    f"INSERT INTO products (name, inventory_number, quantity, unit, min_stock, price, batch) " +
                    f"VALUES ('{safe_name}', '{safe_inv}', {product['quantity']}, '{safe_unit}', " +
                    f"{product['min_stock']}, {product['price']}, '{safe_batch}')"
                )
                inserted += 1
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            'statusCode': 200,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({
                'success': True,
                'inserted': inserted,
                'updated': updated,
                'total': inserted + updated
            })
        }
    
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'isBase64Encoded': False,
            'body': json.dumps({'error': f'Import error: {str(e)}'})
        }